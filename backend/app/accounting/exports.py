from datetime import datetime, timedelta
from typing import List, Optional, Dict, Any, Union
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from app.accounting.models import Obligation, Payment

TEMPLATES_DIR = os.path.join(os.path.dirname(__file__), "templates")
os.makedirs(TEMPLATES_DIR, exist_ok=True)

EXPORTS_DIR = os.path.join(os.path.dirname(__file__), "exports")
os.makedirs(EXPORTS_DIR, exist_ok=True)

def get_template_file(template_name: str) -> Optional[str]:
    """Get the path to a template file."""
    template_path = os.path.join(TEMPLATES_DIR, f"{template_name}_template.pdf")
    if os.path.exists(template_path):
        return template_path
    return None

async def export_obligations_to_excel(
    company_id: Optional[int] = None,
    month: Optional[str] = None,
    status: Optional[str] = None
) -> Optional[str]:
    """Export obligations to Excel file.
    
    Args:
        company_id: Optional filter by company
        month: Optional filter by month (YYYY-MM format)
        status: Optional filter by status
        
    Returns:
        Path to the generated Excel file
    """
    from app.accounting.services import get_obligations, get_company, get_tax_type
    
    filters = {}
    if company_id:
        filters["company_id"] = company_id
    if status:
        filters["status"] = status
    
    obligations = get_obligations(filters=filters)
    
    if month:
        try:
            month_date = datetime.strptime(month, "%Y-%m")
            month_start = month_date.replace(day=1)
            if month_date.month == 12:
                month_end = datetime(month_date.year + 1, 1, 1) - timedelta(days=1)
            else:
                month_end = datetime(month_date.year, month_date.month + 1, 1) - timedelta(days=1)
                
            filtered_obligations = []
            for obligation in obligations:
                if isinstance(obligation.next_due_date, str):
                    due_date = datetime.fromisoformat(obligation.next_due_date.replace("Z", "+00:00"))
                else:
                    due_date = obligation.next_due_date
                    
                if month_start <= due_date <= month_end:
                    filtered_obligations.append(obligation)
            
            obligations = filtered_obligations
        except ValueError:
            pass
    
    if not obligations:
        return None
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Obligations"
    
    headers = [
        "ID", "Company", "Tax Type", "Name", "Status", 
        "Due Date", "Amount", "Frequency"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    for row, obligation in enumerate(obligations, start=2):
        company = get_company(obligation.company_id)
        tax_type = get_tax_type(obligation.tax_type_id)
        
        ws.cell(row=row, column=1, value=obligation.id)
        ws.cell(row=row, column=2, value=company.name if company else "Unknown")
        ws.cell(row=row, column=3, value=tax_type.name if tax_type else "Unknown")
        ws.cell(row=row, column=4, value=obligation.name)
        ws.cell(row=row, column=5, value=obligation.status)
        ws.cell(row=row, column=6, value=obligation.next_due_date.strftime("%Y-%m-%d") if isinstance(obligation.next_due_date, datetime) else obligation.next_due_date)
        ws.cell(row=row, column=7, value=obligation.amount)
        ws.cell(row=row, column=8, value=obligation.frequency)
    
    for col in range(1, len(headers) + 1):
        column_letter = openpyxl.utils.get_column_letter(col)
        max_length = 0
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    company_str = f"_company{company_id}" if company_id else ""
    month_str = f"_month{month}" if month else ""
    status_str = f"_status{status}" if status else ""
    
    file_name = f"obligations{company_str}{month_str}{status_str}_{timestamp}.xlsx"
    file_path = os.path.join(EXPORTS_DIR, file_name)
    wb.save(file_path)
    
    return file_path

async def export_payments_to_excel(
    company_id: Optional[int] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None
) -> Optional[str]:
    """Export payments to Excel file.
    
    Args:
        company_id: Optional filter by company
        from_date: Optional filter by start date (YYYY-MM-DD format)
        to_date: Optional filter by end date (YYYY-MM-DD format)
        
    Returns:
        Path to the generated Excel file
    """
    from app.accounting.services import get_payments, get_obligation, get_company
    
    all_payments = get_payments()
    filtered_payments = []
    
    for payment in all_payments:
        obligation = get_obligation(payment.obligation_id)
        if not obligation:
            continue
            
        if company_id and obligation.company_id != company_id:
            continue
        
        payment_date = payment.payment_date
        if isinstance(payment_date, str):
            payment_date = datetime.fromisoformat(payment_date.replace("Z", "+00:00"))
            
        if from_date:
            try:
                from_datetime = datetime.strptime(from_date, "%Y-%m-%d")
                if payment_date < from_datetime:
                    continue
            except ValueError:
                pass
                
        if to_date:
            try:
                to_datetime = datetime.strptime(to_date, "%Y-%m-%d")
                if payment_date > to_datetime:
                    continue
            except ValueError:
                pass
                
        filtered_payments.append((payment, obligation))
    
    if not filtered_payments:
        return None
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payments"
    
    headers = [
        "ID", "Obligation", "Company", "Amount", 
        "Payment Date", "Receipt Number", "Notes"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    for row, (payment, obligation) in enumerate(filtered_payments, start=2):
        company = get_company(obligation.company_id)
        
        ws.cell(row=row, column=1, value=payment.id)
        ws.cell(row=row, column=2, value=obligation.name)
        ws.cell(row=row, column=3, value=company.name if company else "Unknown")
        ws.cell(row=row, column=4, value=payment.amount)
        ws.cell(row=row, column=5, value=payment.payment_date.strftime("%Y-%m-%d") if isinstance(payment.payment_date, datetime) else payment.payment_date)
        ws.cell(row=row, column=6, value=payment.receipt_number)
        ws.cell(row=row, column=7, value=payment.notes)
    
    for col in range(1, len(headers) + 1):
        column_letter = openpyxl.utils.get_column_letter(col)
        max_length = 0
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    company_str = f"_company{company_id}" if company_id else ""
    from_str = f"_from{from_date}" if from_date else ""
    to_str = f"_to{to_date}" if to_date else ""
    
    file_name = f"payments{company_str}{from_str}{to_str}_{timestamp}.xlsx"
    file_path = os.path.join(EXPORTS_DIR, file_name)
    wb.save(file_path)
    
    return file_path
